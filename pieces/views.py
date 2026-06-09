from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from .models import Piece, MouvementStock
from .forms import PieceForm, ConsommationForm, ReceptionForm
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment
from django.contrib.auth.models import User
from .models import Profile


def dashboard(request):
    pieces = Piece.objects.all()
    total = pieces.count()
    critiques = pieces.filter(criticite=True).count()
    derniers_mouvements = MouvementStock.objects.select_related('piece').all()[:10]
    context = {
        'total_pieces': total,
        'pieces_critiques': critiques,
        'derniers_mouvements': derniers_mouvements,
        'page': 'dashboard',
    }
    return render(request, 'pieces/dashboard.html', context)


def liste_pieces(request):
    query = request.GET.get('q', '')
    emplacement_filtre = request.GET.get('emplacement', '')
    criticite_filtre = request.GET.get('criticite', '')
    pieces = Piece.objects.all()
    if query:
        pieces = pieces.filter(
            Q(description__icontains=query) |
            Q(reference__icontains=query) |
            Q(fournisseur__icontains=query)
        )
    if emplacement_filtre:
        pieces = pieces.filter(emplacement__icontains=emplacement_filtre)
    if criticite_filtre == 'oui':
        pieces = pieces.filter(criticite=True)
    elif criticite_filtre == 'non':
        pieces = pieces.filter(criticite=False)
    context = {
        'pieces': pieces,
        'query': query,
        'emplacement_filtre': emplacement_filtre,
        'criticite_filtre': criticite_filtre,
        'page': 'pieces',
    }
    return render(request, 'pieces/liste_pieces.html', context)


def ajouter_piece(request):
    if request.method == 'POST':
        form = PieceForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pièce ajoutée avec succès.')
            return redirect('liste_pieces')
    else:
        form = PieceForm()
    return render(request, 'pieces/form_piece.html', {'form': form, 'titre': 'Ajouter une pièce', 'page': 'pieces'})


def modifier_piece(request, pk):
    piece = get_object_or_404(Piece, pk=pk)
    if request.method == 'POST':
        form = PieceForm(request.POST, request.FILES, instance=piece)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pièce modifiée avec succès.')
            return redirect('liste_pieces')
    else:
        form = PieceForm(instance=piece)
    return render(request, 'pieces/form_piece.html', {'form': form, 'titre': 'Modifier la pièce', 'piece': piece, 'page': 'pieces'})


def supprimer_piece(request, pk):
    piece = get_object_or_404(Piece, pk=pk)
    if request.method == 'POST':
        piece.delete()
        messages.success(request, 'Pièce supprimée.')
        return redirect('liste_pieces')
    return render(request, 'pieces/confirmer_suppression.html', {'piece': piece, 'page': 'pieces'})


def consommation(request):
    if request.method == 'POST':
        form = ConsommationForm(request.POST)
        if form.is_valid():
            piece = form.cleaned_data['piece']
            quantite = form.cleaned_data['quantite']
            commentaire = form.cleaned_data.get('commentaire', '')
            date_mouvement = form.cleaned_data.get('date') or timezone.now()
            if quantite > piece.stock_reel:
                messages.error(request, f'Stock insuffisant. Stock actuel : {piece.stock_reel}')
            else:
                stock_avant = piece.stock_reel
                piece.stock_reel -= quantite
                piece.save()
                MouvementStock.objects.create(
                    piece=piece,
                    type_mouvement='consommation',
                    quantite=quantite,
                    stock_avant=stock_avant,
                    stock_apres=piece.stock_reel,
                    commentaire=commentaire,
                    matricule=form.cleaned_data.get('matricule', ''),
                    date=date_mouvement,
                )
                messages.success(request, f'Consommation enregistrée. Nouveau stock : {piece.stock_reel}')
                return redirect('consommation')
    else:
        form = ConsommationForm()
    historique = MouvementStock.objects.filter(type_mouvement='consommation').select_related('piece')[:20]
    return render(request, 'pieces/consommation.html', {'form': form, 'historique': historique, 'page': 'consommation'})


def reception(request):
    if request.method == 'POST':
        form = ReceptionForm(request.POST)
        if form.is_valid():
            piece = form.cleaned_data['piece']
            quantite = form.cleaned_data['quantite']
            commentaire = form.cleaned_data.get('commentaire', '')
            date_mouvement = form.cleaned_data.get('date') or timezone.now()
            stock_avant = piece.stock_reel
            piece.stock_reel += quantite
            piece.save()
            MouvementStock.objects.create(
                piece=piece,
                type_mouvement='reception',
                quantite=quantite,
                stock_avant=stock_avant,
                stock_apres=piece.stock_reel,
                commentaire=commentaire,
                date=date_mouvement,
            )
            messages.success(request, f'Réception enregistrée. Nouveau stock : {piece.stock_reel}')
            return redirect('reception')
    else:
        form = ReceptionForm()
    historique = MouvementStock.objects.filter(type_mouvement='reception').select_related('piece')[:20]
    return render(request, 'pieces/reception.html', {'form': form, 'historique': historique, 'page': 'reception'})


def pieces_critiques(request):
    pieces = Piece.objects.all()
    critiques_oui = [p for p in pieces if p.est_critique and p.criticite]
    critiques_non = [p for p in pieces if p.est_critique and not p.criticite]
    a_surveiller = [p for p in pieces if p.statut_stock == 'bas']
    return render(request, 'pieces/pieces_critiques.html', {
        'critiques_oui': critiques_oui,
        'critiques_non': critiques_non,
        'a_surveiller': a_surveiller,
        'page': 'critiques',
    })


def historique(request):
    type_filtre = request.GET.get('type', '')
    mouvements = MouvementStock.objects.select_related('piece').all()
    if type_filtre:
        mouvements = mouvements.filter(type_mouvement=type_filtre)
    return render(request, 'pieces/historique.html', {
        'mouvements': mouvements,
        'type_filtre': type_filtre,
        'page': 'historique',
    })


def export_pieces_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pièces"

    header_fill = PatternFill("solid", fgColor="F5C518")
    header_font = Font(bold=True, color="1A1A2E")

    headers = ['Description', 'Référence', 'Emplacement', 'Stock réel',
               'Fournisseur', 'Stock min', 'Stock max', 'Type', 'Statut', 'Criticité']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for piece in Piece.objects.all():
        ws.append([
            piece.description, piece.reference, piece.emplacement,
            piece.stock_reel, piece.fournisseur, piece.stock_minimum,
            piece.stock_maximum, piece.get_type_piece_display(),
            'Critique' if piece.est_critique else 'Normal',
            'OUI' if piece.criticite else 'NON',
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pieces.xlsx"'
    wb.save(response)
    return response


def export_historique_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historique"

    header_fill = PatternFill("solid", fgColor="F5C518")
    header_font = Font(bold=True, color="1A1A2E")

    headers = ['Date', 'Pièce', 'Référence', 'Type mouvement', 'Matricule',
               'Quantité', 'Stock avant', 'Stock après', 'Commentaire']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for m in MouvementStock.objects.select_related('piece').all():
        ws.append([
            m.date.strftime('%d/%m/%Y %H:%M'),
            m.piece.description, m.piece.reference,
            m.type_mouvement, m.matricule,
            m.quantite, m.stock_avant, m.stock_apres, m.commentaire
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="historique.xlsx"'
    wb.save(response)
    return response


def export_alertes_excel(request):
    wb = openpyxl.Workbook()
    headers = ['Référence', 'Description', 'Emplacement', 'Stock réel',
               'Stock min', 'Fournisseur', 'Type', 'Statut', 'Criticité']

    pieces = list(Piece.objects.all())
    critiques_oui = [p for p in pieces if p.est_critique and p.criticite]
    critiques_non = [p for p in pieces if p.est_critique and not p.criticite]
    a_surveiller = [p for p in pieces if p.statut_stock == 'bas']

    def style_header(ws, fill_color, font_color):
        header_fill = PatternFill("solid", fgColor=fill_color)
        header_font = Font(bold=True, color=font_color)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

    def add_pieces(ws, liste, statut_label):
        for piece in liste:
            ws.append([
                piece.reference, piece.description, piece.emplacement,
                piece.stock_reel, piece.stock_minimum, piece.fournisseur,
                piece.get_type_piece_display(),
                statut_label,
                'OUI' if piece.criticite else 'NON',
            ])

    def auto_width(ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws1 = wb.active
    ws1.title = "Critiques OUI"
    style_header(ws1, "FF0000", "FFFFFF")
    add_pieces(ws1, critiques_oui, 'Critique')
    auto_width(ws1)

    ws2 = wb.create_sheet("Critiques NON")
    style_header(ws2, "FFC107", "000000")
    add_pieces(ws2, critiques_non, 'Critique')
    auto_width(ws2)

    ws3 = wb.create_sheet("A surveiller")
    style_header(ws3, "0D6EFD", "FFFFFF")
    add_pieces(ws3, a_surveiller, 'Stock bas')
    auto_width(ws3)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="alertes_stock.xlsx"'
    wb.save(response)
    return response


def liste_utilisateurs(request):
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        messages.error(request, "Accès réservé aux administrateurs.")
        return redirect('dashboard')
    utilisateurs = User.objects.select_related('profile').all()
    return render(request, 'pieces/utilisateurs/liste.html', {
        'utilisateurs': utilisateurs, 'page': 'utilisateurs'
    })


def ajouter_utilisateur(request):
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        email = request.POST.get('email', '')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        role = request.POST.get('role', 'technicien')
        telephone = request.POST.get('telephone', '')
        poste = request.POST.get('poste', '')

        if User.objects.filter(username=username).exists():
            messages.error(request, "Ce nom d'utilisateur existe déjà.")
        else:
            user = User.objects.create_user(
                username=username, password=password,
                email=email, first_name=first_name, last_name=last_name
            )
            user.profile.role = role
            user.profile.telephone = telephone
            user.profile.poste = poste
            if request.FILES.get('photo'):
                user.profile.photo = request.FILES['photo']
            user.profile.save()
            messages.success(request, f'Utilisateur {username} créé avec succès.')
            return redirect('liste_utilisateurs')
    return render(request, 'pieces/utilisateurs/form.html', {
        'titre': 'Ajouter un utilisateur', 'page': 'utilisateurs'
    })


def modifier_utilisateur(request, pk):
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return redirect('dashboard')
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user.email = request.POST.get('email', '')
        user.first_name = request.POST.get('first_name', '')
        user.last_name = request.POST.get('last_name', '')
        if request.POST.get('password'):
            user.set_password(request.POST.get('password'))
        user.save()
        user.profile.role = request.POST.get('role', 'technicien')
        user.profile.telephone = request.POST.get('telephone', '')
        user.profile.poste = request.POST.get('poste', '')
        if request.FILES.get('photo'):
            user.profile.photo = request.FILES['photo']
        user.profile.save()
        messages.success(request, 'Utilisateur modifié avec succès.')
        return redirect('liste_utilisateurs')
    return render(request, 'pieces/utilisateurs/form.html', {
        'titre': "Modifier l'utilisateur", 'u': user, 'page': 'utilisateurs'
    })


def supprimer_utilisateur(request, pk):
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return redirect('dashboard')
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        if user == request.user:
            messages.error(request, "Vous ne pouvez pas supprimer votre propre compte.")
        else:
            user.delete()
            messages.success(request, 'Utilisateur supprimé.')
        return redirect('liste_utilisateurs')
    return render(request, 'pieces/utilisateurs/confirmer.html', {'u': user, 'page': 'utilisateurs'})


def mon_profil(request):
    if request.method == 'POST':
        request.user.first_name = request.POST.get('first_name', '')
        request.user.last_name = request.POST.get('last_name', '')
        request.user.email = request.POST.get('email', '')
        if request.POST.get('password'):
            request.user.set_password(request.POST.get('password'))
        request.user.save()
        request.user.profile.telephone = request.POST.get('telephone', '')
        request.user.profile.poste = request.POST.get('poste', '')
        if request.FILES.get('photo'):
            request.user.profile.photo = request.FILES['photo']
        request.user.profile.save()
        messages.success(request, 'Profil mis à jour.')
        return redirect('mon_profil')
    return render(request, 'pieces/utilisateurs/profil.html', {'page': 'profil'})